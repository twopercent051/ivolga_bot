from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
import os
from datetime import timedelta, datetime


async def create_xlsx(user_list: list):
    wb = Workbook()
    ws = wb.active
    for user in user_list:
        name = f"{user['name']}"
        add_datetime = (user["add_datetime"] + timedelta(hours=3)).strftime("%d.%m.%Y %H:%M")
        ws.append(
            (
                name,
                user["user_id"],
                add_datetime,
            )
        )

    wb.save(f'{os.getcwd()}/downloaded_db.xlsx')


async def get_xlsx() -> list:
    wb = load_workbook(filename=f"{os.getcwd()}/uploaded.xlsx")
    sh = wb.active
    user_list = []
    for row in sh.iter_rows():
        data = dict(
            name=str(row[0].value),
            user_id=str(row[1].value),
            # add_datetime=datetime.strptime(row[2].value, "%d.%m.%Y %H:%M")
            add_datetime=row[2].value.date(),
            mailing=True
        )
        user_list.append(data)
    return user_list


# async def get_csv():
#     with open(f"{os.getcwd()}/uploaded.csv", newline='') as csvfile:
#         reader = csv.reader(csvfile)
#         user_list = []
#         for row in reader:
#             data = dict(
#                 name=row[0],
#                 user_id=str(row[1]),
#                 add_datetime=datetime.strptime(row[2], "%d.%m.%Y %H:%M")
#             )
#             user_list.append(data)
#         return user_list

